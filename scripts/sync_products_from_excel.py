import json
import re
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
INDEX = ROOT / "index.html"
WORKBOOK = ROOT / "المواصفات الفنية.xlsx"
FALLBACK_WORKBOOK = ROOT / "منتجات الاضاءة MBM LIGHTNIG.xlsx"
SHEET_NAME = "المواصفات الفنية"


C101_FEATURES = "اختيار COB أو SMD؛ عاكس وناشر ضوء اختياريان حسب احتياج الإضاءة."
XFXTD_FEATURES = "جسم نحيف جداً؛ خامة بلاستيكية بالكامل."
TRACKLIGHT_FEATURES = "تراك لايت موجه؛ مناسب للمعارض والمطاعم والمحلات؛ جسم من الألمنيوم والبولي كربونيت."
XYTD_FEATURES = "إضاءة مدمجة مانعة للتوهج؛ مصدر ضوء SMD؛ متوفر باللون الأبيض."
XYTH_FEATURES = "تصميم شبكي Honeycomb؛ حماية مضاعفة من التوهج وراحة أعلى للعين."
MAGNETIC_FEATURES = "نظام مغناطيسي 48V؛ تثبيت قوي؛ قابل للتمديد والتركيب بطرق متعددة."
MAGNETIC_ADJUSTABLE_FEATURES = "نظام مغناطيسي 48V؛ وحدة متحركة قابلة للتوجيه؛ تثبيت قوي وتركيب مرن."
TRACK_ACCESSORY_FEATURES = "ملحقات نظام التراك؛ تساعد على التوصيل والتركيب والتمديد حسب احتياج المشروع."
SUSPENSION_FEATURES = "وحدة تعليق مغناطيسية 48V؛ تثبيت قوي؛ مناسبة للإضاءة الموجهة فوق الطاولات."
DRIVER_FEATURES = "درايفر داخلي لنظام التراك المغناطيسي 48V."


def downlight_specs(angle, cri, size, cutout, ugr=None):
    parts = [f"زاوية: {angle}", f"CRI: {cri}"]
    if ugr:
        parts.append(f"مانع التوهج: UGR{ugr}")
    parts.extend([f"المقاس: {size}", f"قص السقف: {cutout}"])
    return "؛ ".join(parts)


def magnetic_specs(power, color, angle, cri, size):
    return "؛ ".join(
        [
            f"القدرة: {power}",
            f"الألوان: {color}",
            f"زاوية: {angle}",
            f"CRI: {cri}",
            f"المقاس: {size}",
        ]
    )


def clean_text(value):
    if value is None:
        return ""
    return " ".join(str(value).replace("\n", " ").split()).strip()


def parse_fallback_code(code):
    code = clean_text(code)
    match = re.match(r"^(.*?)-(\d+(?:\.\d+)?)W$", code, re.IGNORECASE)
    if match:
        return code, match.group(1).strip(), int(float(match.group(2)))
    return code, code.strip(), None


def normalized_model(model):
    value = clean_text(model).upper()
    value = value.replace("CXZD", "CXD")
    return re.sub(r"[^A-Z0-9/]", "", value)


def format_numeric(value):
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return clean_text(value)


def fallback_entry_to_specs(entry):
    parts = []
    if entry.get("material") and entry["material"] != "/":
        parts.append(f"الخامة: {entry['material']}")
    if entry.get("cri") and entry["cri"] != "/":
        parts.append(f"CRI: {entry['cri']}")
    if entry.get("led_type") and entry["led_type"] != "/":
        parts.append(f"نوع الإضاءة: {entry['led_type']}")
    if entry.get("ip") and entry["ip"] != "/":
        parts.append(f"الحماية: {entry['ip']}")
    if entry.get("ccts"):
        parts.append(f"الألوان: {'/'.join(entry['ccts'])}")
    return "؛ ".join(parts) if parts else "-"


def load_fallback_specs():
    if not FALLBACK_WORKBOOK.exists():
        return {}

    workbook = openpyxl.load_workbook(FALLBACK_WORKBOOK, data_only=True)
    sheet = workbook.active
    entries = {}

    for row in sheet.iter_rows(min_row=2, max_col=10, values_only=True):
        code = clean_text(row[1])
        if not code:
            continue
        raw_code, model, watt = parse_fallback_code(code)
        key = (normalized_model(model), watt)
        entry = entries.setdefault(
            key,
            {
                "raw_code": raw_code,
                "model": model,
                "watt": watt,
                "material": "",
                "cri": "",
                "led_type": "",
                "ip": "",
                "ccts": [],
            },
        )
        material = clean_text(row[5])
        cri = format_numeric(row[6])
        led_type = clean_text(row[7])
        ip = clean_text(row[8])
        cct = clean_text(row[9])
        if material and material != "/" and not entry["material"]:
            entry["material"] = material
        if cri and cri != "/" and not entry["cri"]:
            entry["cri"] = cri
        if led_type and led_type != "/" and not entry["led_type"]:
            entry["led_type"] = led_type
        if ip and ip != "/" and not entry["ip"]:
            entry["ip"] = ip
        if cct and cct != "/" and cct not in entry["ccts"]:
            entry["ccts"].append(cct)

    return entries


def fallback_entry_for(product, fallback_specs):
    model_key = normalized_model(product["model"])
    exact = fallback_specs.get((model_key, product["watt"]))
    if exact:
        return exact

    if model_key.startswith("XYTD") and product["watt"] is not None:
        same_family_watt = [
            entry
            for (model, watt), entry in fallback_specs.items()
            if model.startswith("XYTD") and watt == product["watt"]
        ]
        if same_family_watt:
            return same_family_watt[0]

    same_model = [
        entry for (model, _), entry in fallback_specs.items() if model == model_key
    ]
    if same_model:
        return same_model[0]

    fuzzy = []
    for (model, watt), entry in fallback_specs.items():
        same_watt = watt is None or product["watt"] is None or watt == product["watt"]
        if not same_watt:
            continue
        if model_key and (model.startswith(model_key[:-1]) or model_key.startswith(model[:-1])):
            fuzzy.append(entry)
    return fuzzy[0] if fuzzy else None


def specs_from_accessory(product):
    name = product["item_name"]
    watt = product["watt"]
    if "درايفر" in name or "محول طاقة" in name:
        parts = ["النظام: 48V"]
        if watt:
            parts.append(f"القدرة: {watt}W")
        return "؛ ".join(parts)
    if "تراك" in name:
        length_match = re.search(r"(\d+(?:\.\d+)?)\s*متر", name)
        track_type = "مخفي" if "مخفي" in name else "معلق" if "معلق" in name else "تراك"
        parts = [f"النوع: تراك {track_type}".replace("تراك تراك", "تراك")]
        if length_match:
            parts.append(f"الطول: {length_match.group(1)} متر")
        return "؛ ".join(parts)
    if "وصلة" in name or "توصيلة" in name:
        return "النوع: وصلة تراك مغناطيسي؛ الاستخدام: ربط وتمديد المسار"
    if "مشبك" in name:
        return "النوع: مشبك تثبيت؛ الاستخدام: تثبيت التراك المغناطيسي"
    if "صفيحة" in name:
        return "النوع: صفيحة تثبيت؛ الاستخدام: دعم تثبيت التراك"
    if "سلك" in name:
        return "النوع: سلك توصيل؛ الاستخدام: تغذية وربط نظام التراك"
    return "-"


def family_features(product):
    model = product["model"]
    name = product["item_name"]
    if model.startswith("XFX-TD"):
        return XFXTD_FEATURES
    if model.startswith("GY-TL"):
        return TRACKLIGHT_FEATURES
    if model.startswith("XY-TD"):
        return XYTD_FEATURES
    if model.startswith("XYTH"):
        return XYTH_FEATURES
    if "متحرك" in name:
        return MAGNETIC_ADJUSTABLE_FEATURES
    if model.startswith("CXDX"):
        return SUSPENSION_FEATURES
    if model.startswith("CX"):
        return MAGNETIC_ADJUSTABLE_FEATURES
    if "درايفر" in name or "محول طاقة" in name:
        return DRIVER_FEATURES
    if model == "N/A" or "تراك" in name or "وصلة" in name or "توصيلة" in name:
        return TRACK_ACCESSORY_FEATURES
    return "-"


BROCHURE_BY_MODEL = {
    "CY107": {
        "lumen": "630",
        "beam_angle": "90°",
        "cri": "≥90",
        "technical_specs": downlight_specs("90°", "≥90", "Φ88*40", "Φ75"),
        "features": C101_FEATURES,
    },
    "CY112": {
        "lumen": "1000",
        "beam_angle": "90°",
        "cri": "≥90",
        "technical_specs": downlight_specs("90°", "≥90", "Φ120*45", "Φ100"),
        "features": C101_FEATURES,
    },
    "CY120": {
        "lumen": "1800",
        "beam_angle": "90°",
        "cri": "≥90",
        "technical_specs": downlight_specs("90°", "≥90", "Φ166*45", "Φ150"),
        "features": C101_FEATURES,
    },
    "CY130": {
        "lumen": "2700",
        "beam_angle": "90°",
        "cri": "≥90",
        "technical_specs": downlight_specs("90°", "≥90", "Φ223*46", "Φ200"),
        "features": C101_FEATURES,
    },
    "CY140": {
        "lumen": "3600",
        "beam_angle": "90°",
        "cri": "≥90",
        "technical_specs": downlight_specs("90°", "≥90", "Φ223*46", "Φ200"),
        "features": C101_FEATURES,
    },
    "XFX-TD1203": {
        "lumen": "950",
        "beam_angle": "110°",
        "cri": "≥80",
        "technical_specs": downlight_specs("110°", "≥80", "Φ138*32", "Φ125"),
        "features": XFXTD_FEATURES,
    },
    "XYTH403": {
        "lumen": "240",
        "beam_angle": "20°",
        "cri": "≥90",
        "ugr": "<16",
        "technical_specs": downlight_specs("20°", "≥90", "Φ69*45", "Φ55-60", "<16"),
        "features": XYTH_FEATURES,
    },
    "XYTH703": {
        "lumen": "490",
        "beam_angle": "20°",
        "cri": "≥90",
        "ugr": "<16",
        "technical_specs": downlight_specs("20°", "≥90", "Φ85*56", "Φ75", "<16"),
        "features": XYTH_FEATURES,
    },
    "XYTH1203": {
        "lumen": "840",
        "beam_angle": "20°",
        "cri": "≥90",
        "ugr": "<16",
        "technical_specs": downlight_specs("20°", "≥90", "Φ110*70", "Φ95-100", "<16"),
        "features": XYTH_FEATURES,
    },
    "CXFG20/300": {
        "beam_angle": "120°",
        "cri": "Ra≥90",
        "ugr": ">15",
        "technical_specs": magnetic_specs("10W", "3000K/4000K/6000K", "120°", "Ra≥90", "L300*W22*H44mm"),
        "features": MAGNETIC_FEATURES,
    },
    "CXFG20/600": {
        "beam_angle": "120°",
        "cri": "Ra≥90",
        "ugr": ">15",
        "technical_specs": magnetic_specs("20W", "3000K/4000K/6000K", "120°", "Ra≥90", "L600*W22*H44mm"),
        "features": MAGNETIC_FEATURES,
    },
    "CXDX20/10": {
        "beam_angle": "24°",
        "cri": "Ra≥90",
        "ugr": ">15",
        "technical_specs": magnetic_specs("6W/10W", "3000K/4000K", "24°", "Ra≥90", "Φ30*H300mm"),
        "features": SUSPENSION_FEATURES,
    },
}


BROCHURE_BY_MODEL_WATT = {
    ("XY-TD803", 8): {
        "lumen": "600",
        "beam_angle": "70°",
        "cri": "≥90",
        "ugr": "<19",
        "technical_specs": downlight_specs("70°", "≥90", "Φ88*46", "Φ75", "<19"),
        "features": XYTD_FEATURES,
    },
    ("XY-TD1503", 15): {
        "lumen": "1125",
        "beam_angle": "75°",
        "cri": "≥90",
        "ugr": "<19",
        "technical_specs": downlight_specs("75°", "≥90", "Φ110*50", "Φ95", "<19"),
        "features": XYTD_FEATURES,
    },
    ("XY-TD2403", 24): {
        "lumen": "2190/2600",
        "beam_angle": "60°",
        "cri": "≥90",
        "ugr": "<19",
        "technical_specs": downlight_specs("60°", "≥90", "Φ160*58", "Φ145", "<19"),
        "features": XYTD_FEATURES,
    },
    ("CXZD20/12", 12): {
        "beam_angle": "24°",
        "cri": "Ra≥90",
        "ugr": ">15",
        "technical_specs": magnetic_specs("12W", "3000K/4000K", "24°", "Ra≥90", "L220*W22*H105mm"),
        "features": MAGNETIC_FEATURES,
    },
    ("CXZD20/18", 18): {
        "beam_angle": "24°",
        "cri": "Ra≥90",
        "ugr": ">15",
        "technical_specs": magnetic_specs("18W", "3000K/4000K", "24°", "Ra≥90", "L330*W22*H105mm"),
        "features": MAGNETIC_FEATURES,
    },
}


BROCHURE_BY_ITEM_NUMBER = {
    "07-001-18-015": {
        "technical_specs": "النظام: 48V؛ القدرة: 100W؛ المقاس: L230*W22*H43mm",
        "features": DRIVER_FEATURES,
    },
    "07-001-18-016": {
        "technical_specs": "النظام: 48V؛ القدرة: 200W؛ المقاس: L350*W22*H43mm",
        "features": DRIVER_FEATURES,
    },
}


def brochure_spec_for(item_number, model, watt):
    return (
        BROCHURE_BY_ITEM_NUMBER.get(item_number)
        or BROCHURE_BY_MODEL_WATT.get((model, watt))
        or BROCHURE_BY_MODEL.get(model)
        or {}
    )


def parse_watt(value):
    if value is None:
        return None
    match = re.search(r"\d+(?:\.\d+)?", str(value))
    if not match:
        return None
    number = float(match.group(0))
    return int(number) if number.is_integer() else number


def icon_for(name, watt):
    if watt is None:
        return "🔗"
    if "تراك" in name:
        return "🎥"
    if "ماجناتك" in name or "مغناط" in name:
        return "⬛"
    if "بنل" in name:
        return "◻️"
    if "سبوت" in name or "توهج" in name:
        return "💡"
    return "🔲"


def load_products():
    workbook = openpyxl.load_workbook(WORKBOOK, data_only=True)
    sheet = workbook[SHEET_NAME]
    fallback_specs = load_fallback_specs()
    products = []

    for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), 1):
        item_number, item_name, model, watt_text, kelvin = row
        if not item_number:
            continue

        item_name = str(item_name).strip() if item_name is not None else ""
        model = str(model).strip() if model is not None else "N/A"
        watt_text = str(watt_text).strip() if watt_text is not None else ""
        kelvin = str(kelvin).strip() if kelvin is not None else None
        watt = parse_watt(watt_text)
        item_number = str(item_number).strip()
        brochure_spec = brochure_spec_for(item_number, model or "N/A", watt)

        product = {
            "id": idx,
            "item_number": item_number,
            "item_name": item_name,
            "model": model or "N/A",
            "category": "Excel Catalog",
            "watt": watt,
            "watt_text": watt_text,
            "kelvin": kelvin,
            "lumen": "-",
            "cri": "-",
            "beam_angle": "-",
            "ugr": "-",
            "technical_specs": "-",
            "features": "-",
            "icon": icon_for(item_name, watt),
        }
        product.update(brochure_spec)

        if product["technical_specs"] == "-":
            fallback_entry = fallback_entry_for(product, fallback_specs)
            if fallback_entry:
                product["technical_specs"] = fallback_entry_to_specs(fallback_entry)
            if product["technical_specs"] == "-":
                product["technical_specs"] = specs_from_accessory(product)

        if product["features"] == "-":
            product["features"] = family_features(product)

        products.append(product)

    features_by_model = {}
    for product in products:
        if product["model"] != "N/A" and product["features"] != "-":
            features_by_model.setdefault(product["model"], product["features"])

    for product in products:
        if product["features"] == "-" and product["model"] in features_by_model:
            product["features"] = features_by_model[product["model"]]

    return products


def to_js_value(value, indent=6):
    text = json.dumps(value, ensure_ascii=False, indent=2)
    return "\n".join((" " * indent) + line for line in text.splitlines())


def main():
    html = INDEX.read_text(encoding="utf-8")
    products = load_products()
    products_js = to_js_value(products)
    replacement = (
        "    // --- MBM product data from Excel sheet (المواصفات الفنية) ---\n"
        "    // Source columns: رقم الصنف، اسم الصنف، الموديل، الوات، اللون.\n"
        f"    const mbmProducts = {products_js.strip()};"
    )

    updated = re.sub(
        r"    // --- MBM (?:Real Data Structure mapped with Excel \(المواصفات الفنية\)|product data from Excel sheet \(المواصفات الفنية\)) ---[\s\S]*?    \];",
        replacement,
        html,
        count=1,
    )

    if updated == html:
        raise RuntimeError("Could not find mbmProducts block in index.html")

    INDEX.write_text(updated, encoding="utf-8")
    print(f"Synced {len(products)} products from {SHEET_NAME}.")


if __name__ == "__main__":
    main()
